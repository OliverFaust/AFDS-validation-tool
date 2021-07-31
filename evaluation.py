"""
Copyright (c) 01.07.2021, Dr Dr Oliver Faust

This program is free software: you can redistribute it and/or modify it
under the terms of the GNU General Public License (GPL) as published by
the Free Software Foundation, either version 3 of this License,
or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but without any warranty, without even the implied warranty of
merchantibility or fitness for a particular purpose.
See the GNU General Public License for more details.

You should have received a copy of the GNU GPL along with this program.
If not, please see <https://www.gnu.org/licenses/#GPL>.
"""

from PyQt5.QtCore import QRunnable, QMetaObject, Qt, Q_ARG
import openpyxl
import excelProcessing
from datetime import datetime
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from sklearn.metrics import classification_report
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
import visualisation_utils as my_vis
from PyQt5.QtGui import QPixmap
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve, auc
import numpy as np

class Evaluation(QRunnable):
    def __init__(self, main, rrStartEndList, ecgStartEndList):
        QRunnable.__init__(self)
        self.main = main
        self.rrStartEndList = rrStartEndList
        self.ecgStartEndList = ecgStartEndList
        
    def timeInRange(self, start, end, x):
        if start <= end:
            return start <= x <= end
        else:
            return start <= x or x <= end

    def run(self):
        ecgLabelVec = []
        rrLabelVec = []
        tsVec, rrVec = excelProcessing.getTsRrVEC(self.main.workbook)
        for ts in tsVec:
            timestamp = datetime.strptime(ts,
                                          '%Y-%m-%d %H:%M:%S.%f')
            ecgLabel = 0
            for ecgRegion in self.ecgStartEndList:
                if self.timeInRange(ecgRegion[0], ecgRegion[1], timestamp):
                    ecgLabel = 1
                    break
            ecgLabelVec.append(ecgLabel)
            rrLabel = 0
            for rrRegion in self.rrStartEndList:
                if self.timeInRange(rrRegion[0], rrRegion[1], timestamp):
                    rrLabel = 1
                    break
            rrLabelVec.append(rrLabel)
        classes = ['normal', 'af']
        cm = confusion_matrix(ecgLabelVec, rrLabelVec)

        plt.figure(figsize=[5,5])
        my_vis.plot_confusion_matrix(cm, 
                              classes=classes,
                              title=None)
        fileName = self.main.sourceFileName.replace('.xlsx', '_CM.png')
        plt.savefig(fileName,
                    dpi=100, bbox_inches='tight', pad_inches=0.5)

        tsVec, estAfVec = excelProcessing.getTsEstAfVEC(self.main.workbook)
        fpr, tpr, thresholds = roc_curve(ecgLabelVec[100:], estAfVec)
        roc_auc = auc(fpr, tpr)
        
        plt.figure(figsize=[5,5])
        lw = 2
        plt.plot(fpr, tpr, color='darkorange',
                 lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
        plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.title('Receiver operating characteristic example')
        plt.legend(loc="lower right")
        
        fileNameRoc = self.main.sourceFileName.replace('.xlsx', '_ROC.png')
        plt.savefig(fileNameRoc,
                    dpi=100, bbox_inches='tight', pad_inches=0.5)
        
        # accuracy
        accuracy = accuracy_score(ecgLabelVec, rrLabelVec)
        print('blind-fold accuracy is {0:.5f}'.format(accuracy))
        
        # precision
        precision = precision_score(ecgLabelVec, rrLabelVec)
        print('blind-fold precision is {0:.5f}'.format(precision))
        
        # recall
        recall = recall_score(ecgLabelVec, rrLabelVec)
        print('blind-fold recall is {0:.5f}'.format(recall))
        
        # f1 score
        f1 = f1_score(ecgLabelVec, rrLabelVec)
        print('blind-fold f1 score is {0:.5f}'.format(f1))
        
        try:
            # classification report
            print(classification_report(ecgLabelVec, rrLabelVec, 
                                    target_names=['normal', 'af']))    
        
            TP=cm[1,1]
            TN=cm[0,0]
            FP=cm[0,1]
            FN=cm[1,0]
            
            ACC=(TP+TN)/(TP+FP+FN+TN)
            SEN=(TP)/(TP+FN)
            SPE=(TN)/(TN+FP)
            if 'AF results' not in self.main.workbook.sheetnames:
                self.main.workbook.create_sheet('AF results')
            sheet = excelProcessing.getSheet(self.main.workbook, 'AF results')
            sheet['A1'] = 'Confusion Matrix'
            sheet['A2'] = 'Normal'
            sheet['A3'] = 'AF'
            sheet['B2'] = TN
            sheet['B3'] = FN
            sheet['C2'] = FP
            sheet['C3'] = TP        
            sheet['A5'] = 'Accuracy'
            sheet['B5'] = ACC
            sheet['A6'] = 'Sensitivity'
            sheet['B6'] = SEN
            sheet['A7'] = 'Specificity'
            sheet['B7'] = SPE
            sheet['A8'] = 'F1-score'  
            sheet['B8'] = f1
            img = openpyxl.drawing.image.Image(fileName)
            img.anchor = 'E1'
            sheet.add_image(img)
            
            if 'ROC' not in self.main.workbook.sheetnames:
                self.main.workbook.create_sheet('ROC')
            sheet = excelProcessing.getSheet(self.main.workbook, 'ROC')
            sheet['A1'] = 'Receiver operating characteristic'
            sheet['A2'] = 'AUC'
            sheet['A3'] = 'Threshold'
            sheet['B2'] = roc_auc
            sheet['B3'] = thresholds[np.argmax(tpr - fpr)]       
            img = openpyxl.drawing.image.Image(fileNameRoc)
            img.anchor = 'E1'
            sheet.add_image(img)

        except:
            print("An exception occured during evaluation") 

        QMetaObject.invokeMethod(self.main, "setEvaluated",
                                 Qt.QueuedConnection)
