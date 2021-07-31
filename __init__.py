"""
Copyright (c) 01.07.2021, Dr Dr Oliver Faust

This program is free software: you can redistribute it and/or modify it
under the terms of the GNU General Public License (GPL) as published by
the Free Software Foundation, either version 3 of this License,
or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but without any warranty, without even the implied warranty of
merchantibility or fitness for a particular purpose.
See the GNU General Public License for more details.

You should have received a copy of the GNU GPL along with this program.
If not, please see <https://www.gnu.org/licenses/#GPL>.
"""

from PyQt5 import QtWidgets, uic
from PyQt5.QtCore import Qt, QThreadPool, pyqtSlot
from PyQt5.QtWidgets import QFileDialog
import pyqtgraph as pg
import sys
import excelProcessing
import DateAxisItem
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from datetime import datetime
from WaitingSpinnerWidget import QtWaitingSpinner
from plotProcessing import PlotProcessing
from ecgPlotProcessing import EcgPlotProcessing
from estimateAfProbability import EstimateAfProbability
from loadExcel import LoadExcel
from enum import Enum
from PyQt5.QtGui import QPixmap
from regionManagement import RegionManagement
from evaluation import Evaluation
from saveWorkbook import SaveWorkbook
 
class state(Enum):
    AF = 1
    nonAF = 0
        
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        uic.loadUi('mainWindow_v3.ui', self)
        self.workbook = None
        self.sourceFileName = []
        self.ecgAfRegions = RegionManagement()
        self.rrAfRegions = RegionManagement()
        self.raiseButton = QtWidgets.QPushButton(self)
        self.raiseButton.setFixedHeight(0)
        self.setCentralWidget(self.raiseButton)
        self.dockList = []
        self.addDockWidget(Qt.TopDockWidgetArea, self.controlDock)
        self.addDockWidget(Qt.TopDockWidgetArea, self.graphsDock)
        self.dockList.insert(0, self.controlDock)
        self.dockList.insert(1, self.graphsDock)
        self.dockList.insert(2, self.configDock)
        self.dockList.insert(3, self.ecgRegions)
        self.dockList.insert(4, self.inferencePerformance)
        self.dockList.insert(5, self.rocDock)
        if len(self.dockList) > 1:
            for index in range(0, len(self.dockList) - 1):
                self.tabifyDockWidget(self.dockList[index],
                                      self.dockList[index + 1])
        self.controlDock.setWindowTitle('Control')
        self.graphsDock.setWindowTitle('Graphs')
        self.configDock.setWindowTitle('RR AF region config')
        self.ecgRegions.setWindowTitle('ECG AF region config')
        self.inferencePerformance.setWindowTitle('Inference Performance')
        self.rocDock.setWindowTitle('ROC')
        self.dockList[0].raise_()
        self.labelAllAfRegionsButton.clicked.connect(self.labelAllAfRegions)
        self.loadButton.clicked.connect(self.loadFile)
        self.saveButton.clicked.connect(self.saveFile)
        self.plotButton.clicked.connect(self.plotData)
        self.ecgPlotButton.clicked.connect(self.plotEcg)
        self.addAfButton.clicked.connect(self.addAf)
        self.removeAfButton.clicked.connect(self.removeAf)
        self.addEcgAfButton.clicked.connect(self.ecgAddAf)
        self.removeEcgAfButton.clicked.connect(self.ecgRemoveAf)
        self.processButton.clicked.connect(self.processData)
        self.resultAnalysisButton.clicked.connect(self.resultAnalysis)
        self.threshold.valueChanged.connect(self.thresholdChange)
        axis4 = DateAxisItem.DateAxisItem(orientation='bottom')
        axis4.attachToPlotItem(self.graphWidget4.getPlotItem())
        self.graphWidget4.showGrid(True, True, 0.5)
        axis1 = DateAxisItem.DateAxisItem(orientation='bottom')
        axis1.attachToPlotItem(self.graphWidget1.getPlotItem())
        self.graphWidget1.showGrid(True, True, 0.5)
        axis2 = DateAxisItem.DateAxisItem(orientation='bottom')
        axis2.attachToPlotItem(self.graphWidget2.getPlotItem())
        self.graphWidget2.showGrid(True, True, 0.5)
        axis3 = DateAxisItem.DateAxisItem(orientation='bottom')
        axis3.attachToPlotItem(self.graphWidget3.getPlotItem())
        self.graphWidget3.showGrid(True, True, 0.5)
        self.graphWidget4.setLabel('left', "Amp/mV")
        self.graphWidget1.setLabel('left', "RR/ms")
        self.graphWidget2.setLabel('left', "EstAfP")
        self.graphWidget3.setLabel('left', "EstAfP")
        self.graphWidget3.setLabel('bottom', "Time/s")
        self.spinner = QtWaitingSpinner(self)
        self.layout().addWidget(self.spinner)
    
    def timeInRange(self, start, end, x):
        if start <= end:
            return start <= x <= end
        else:
            return start <= x or x <= end
        
    def resultAnalysis(self):
        self.spinner.start()
        rrStartEndList = self.rrAfRegions.getStartEndList(self.rrAfRegionTable)
        ecgStartEndList = self.ecgAfRegions.getStartEndList(self.ecgAfRegionTable)
        runnable = Evaluation(self, rrStartEndList, ecgStartEndList)
        QThreadPool.globalInstance().start(runnable)

    @pyqtSlot()
    def setEvaluated(self):
        self.spinner.stop()
        sheet = excelProcessing.getSheet(self.workbook, 'AF results')
        if sheet is not None:
            fileName = self.sourceFileName.replace('.xlsx', '_CM.png')
            pix = QPixmap(fileName)
            self.cmView.setPixmap(pix)
            ACC = sheet['B5'].value 
            SEN = sheet['B6'].value
            SPE = sheet['B7'].value 
            f1 = sheet['B8'].value
            self.accuracyLabel.setText('Accuracy = %.4f' % ACC)
            self.sensitivityLabel.setText('Sensitivity = %.4f' % SEN)
            self.specificityLabel.setText('Specificity = %.4f' % SPE)
            self.f1Label.setText('F1-score = %.4f' % f1)    
        sheet = excelProcessing.getSheet(self.workbook, 'ROC')
        if sheet is not None:            
            fileName = self.sourceFileName.replace('.xlsx', '_ROC.png')
            pix = QPixmap(fileName)  
            self.rocView.setPixmap(pix) 
            AUC = sheet['B2'].value 
            TH = sheet['B3'].value
            self.aucLabel.setText('AUC = %.4f' % AUC)
            self.thresholdLabel.setText('Threshold = %.4f' % TH)            
            
    def labelAllAfRegions(self):
        currentState = state.nonAF
        color = (255, 0, 0, 50)
        tsVec, estAfVec = excelProcessing.getTsEstAfVEC(self.workbook)
        for ts, estAf in zip(tsVec, estAfVec):
            if currentState == state.nonAF:
                if estAf > self.threshold.value():    
                    currentState = state.AF
                    time = datetime.strptime(ts,
                                  '%Y-%m-%d %H:%M:%S.%f')
                    startTime = time.timestamp() - 0.05
                    time = datetime.strptime(ts,
                                  '%Y-%m-%d %H:%M:%S.%f')                            
                    endTime = time.timestamp() + 0.05                      
            else:
                if estAf < self.threshold.value():    
                    currentState = state.nonAF;
                    self.rrAfRegions.addRegion(self.rrAfRegionTable, 
                                               self.graphWidget1, 
                                               self.onAfChange, startTime, 
                                               endTime, color)                    
                else:
                    time = datetime.strptime(ts,
                                  '%Y-%m-%d %H:%M:%S.%f')                            
                    endTime = time.timestamp() + 0.05    

    def ecgRemoveAf(self):
        self.ecgAfRegions.remove(self.ecgAfRegionTable, self.graphWidget4)

    def ecgAddAf(self):
        color = (255, 255, 0, 50)
        pos = self.region.getRegion()
        self.ecgAfRegions.addDefaultRegion(self.ecgAfRegionTable, self.graphWidget4, 
                             self.onEcgAfChange, pos, color)

    def removeAf(self):
        self.rrAfRegions.remove(self.rrAfRegionTable, self.graphWidget1)

    def addAf(self):
        color = (255, 0, 0, 50)
        pos = self.region.getRegion()
        self.rrAfRegions.addDefaultRegion(self.rrAfRegionTable, self.graphWidget1, 
                             self.onAfChange, pos, color)

    def loadFile(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        self.sourceFileName, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "", "Excel files (*.xlsx);;All Files (*)", options=options)
        if self.sourceFileName:
            self.plotButton.setEnabled(False)
            self.processButton.setEnabled(False)
            self.addAfButton.setEnabled(False)
            self.removeAfButton.setEnabled(False)
            self.addEcgAfButton.setEnabled(False)
            self.removeEcgAfButton.setEnabled(False)
            self.processButton.setEnabled(False)
            self.resultAnalysisButton.setEnabled(False)
            self.labelAllAfRegionsButton.setEnabled(False)
            self.ecgPlotButton.setEnabled(False)
            self.rrAfRegionTable.clear()
            self.ecgAfRegionTable.clear()
            self.spinner.start()
            runnable = LoadExcel(self, self.sourceFileName)
            QThreadPool.globalInstance().start(runnable)
        else:
            self.loadFilePath.setText('Not yet loaded')

    @pyqtSlot(openpyxl.workbook.workbook.Workbook)
    def setData(self, data):
        self.workbook = data
        self.spinner.stop()
        sheet = excelProcessing.getSheet(self.workbook, 'AF results')
        if sheet is None:
            self.accuracyLabel.setText('Accuracy = None')
            self.sensitivityLabel.setText('Sensitivity = None')
            self.specificityLabel.setText('Specificity = None')
            self.f1Label.setText('F1-score = None')      
            self.cmView.clear()
        else:
            image_loader = SheetImageLoader(sheet)
            image = image_loader.get('E1')
            fileName = self.sourceFileName.replace('.xlsx', '_CM.png')
            image.save(fileName)
            pix = QPixmap(fileName)
            self.cmView.setPixmap(pix)
            ACC = sheet['B5'].value 
            SEN = sheet['B6'].value
            SPE = sheet['B7'].value 
            f1 = sheet['B8'].value
            if ACC is not None:
                self.accuracyLabel.setText('Accuracy = %.4f' % ACC)
            else:
                self.accuracyLabel.setText('Accuracy = None')
            if SEN is not None:
                self.sensitivityLabel.setText('Sensitivity = %.4f' % SEN)
            else:
                self.sensitivityLabel.setText('Sensitivity = None')
            if SPE is not None:
                self.specificityLabel.setText('Specificity = %.4f' % SPE)
            else:
                self.specificityLabel.setText('Specificity = None')
            if f1 is not None:
                self.f1Label.setText('F1-score = %.4f' % f1)
            else:
                self.f1Label.setText('F1-score = None')
                
        sheet = excelProcessing.getSheet(self.workbook, 'ROC')
        if sheet is None:
            self.aucLabel.setText('AUC = None')
            self.thresholdLabel.setText('Threshold = None')
            self.rocView.clear()
        else:
            image_loader = SheetImageLoader(sheet)
            image = image_loader.get('E1')
            fileName = self.sourceFileName.replace('.xlsx', '_ROC.png')
            image.save(fileName)
            pix = QPixmap(fileName)
            self.rocView.setPixmap(pix)
            AUC = sheet['B2'].value 
            TH = sheet['B3'].value
            if AUC is not None:
                self.aucLabel.setText('AUC = %.4f' % AUC)
            else:
                self.aucLabel.setText('AUC = None')
            if SEN is not None:
                self.thresholdLabel.setText('Sensitivity = %.4f' % TH)
            else:
                self.thresholdLabel.setText('Sensitivity = None')

    def thresholdChange(self):
        self.threesholdLabel.setText('Threshold value = %.2f' % self.threshold.value())

    def saveFile(self):
        suggestedFileName = self.sourceFileName.replace('.xlsx', '_AF.xlsx')
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self, "QFileDialog.getSaveFileName()", suggestedFileName, "All Files (*);;Text Files (*.txt)", options=options)      
        self.spinner.start()
        runnable = SaveWorkbook(self, fileName)
        QThreadPool.globalInstance().start(runnable)

    @pyqtSlot()
    def setSaveWorkbook(self):
        self.spinner.stop()
                
    def processData(self):
        self.spinner.start()
        runnable = EstimateAfProbability(self, self.workbook)
        QThreadPool.globalInstance().start(runnable)

    @pyqtSlot(openpyxl.workbook.workbook.Workbook)
    def setProcessedWorkbook(self, workbook):
        self.workbook = workbook
        self.saveButton.setEnabled(True)
        self.plotButton.setEnabled(True)
        self.resultAnalysisButton.setEnabled(True)
        self.spinner.stop()

    def updateRegion(self, window, viewRange):
        rgn = viewRange[0]
        self.region.setRegion(rgn)

    def update(self):
        self.region.setZValue(10)
        minX, maxX = self.region.getRegion()
        self.graphWidget1.setXRange(minX, maxX, padding=0)
        self.graphWidget2.setXRange(minX, maxX, padding=0)
        self.graphWidget4.setXRange(minX, maxX, padding=0)

    def mouseMoved1(self, evt):
        if self.graphWidget1.sceneBoundingRect().contains(evt):
            mousePoint = self.vb1.mapSceneToView(evt)
            self.updateMeasurement(mousePoint.x())
        self.ecgAfRegionTable.clearSelection()
        self.rrAfRegions.highlightRow(self.rrAfRegionTable, evt)

    def mouseMoved2(self, evt):
        if self.graphWidget2.sceneBoundingRect().contains(evt):
            mousePoint = self.vb2.mapSceneToView(evt)
            self.updateMeasurement(mousePoint.x())

    def mouseMoved4(self, evt):
        if self.graphWidget4.sceneBoundingRect().contains(evt):
            mousePoint = self.vb4.mapSceneToView(evt)
            self.updateMeasurement(mousePoint.x())
        self.rrAfRegionTable.clearSelection()
        self.ecgAfRegions.highlightRow(self.ecgAfRegionTable, evt)

    def updateMeasurement(self, x):
        date_time = datetime.fromtimestamp(x)
        timeStr = datetime.strftime(date_time, '%Y-%m-%d %H:%M:%S.%f')
        self.markerTime.setText("Marker time: "+timeStr)
        self.ecgMarkerTime.setText("Marker time: "+timeStr)
        self.vLine4.setPos(x)
        self.vLine1.setPos(x)
        self.vLine2.setPos(x)
        
    def plotEcg(self):
        startEnd = self.rrAfRegions.getSequencedStartEndList(self.rrAfRegionTable)
        startEnd.extend( self.ecgAfRegions.getSequencedStartEndList(self.ecgAfRegionTable))      
        self.spinner.start()
        runnable = EcgPlotProcessing(self, self.workbook, startEnd)
        QThreadPool.globalInstance().start(runnable)

    @pyqtSlot(list)
    def setEcgPlotProcessed(self, ecgSnippets):
        self.spinner.stop()
        for ecgSnippet in ecgSnippets:
            self.graphWidget4.plot(ecgSnippet["exgTimestampVec"], ecgSnippet["ecgVec"], pen='w')
        
    def plotData(self):
        self.graphWidget1.clear()
        self.graphWidget2.clear()
        self.graphWidget3.clear()
        self.graphWidget4.clear()
        self.spinner.start()
        runnable = PlotProcessing(self, self.workbook)
        QThreadPool.globalInstance().start(runnable)

    @pyqtSlot(dict)
    def setPlotProcessed(self, data):
        self.spinner.stop()
        sheet = excelProcessing.getSheet(self.workbook, 'AF results')
        if sheet is not None:
            image_loader = SheetImageLoader(sheet)
            image = image_loader.get('E1')
            fileName = self.sourceFileName.replace('.xlsx', '_CM.png')
            image.save(fileName)
            pix = QPixmap(fileName)
            self.cmView.setPixmap(pix)
            ACC = sheet['B5'].value 
            SEN = sheet['B6'].value
            SPE = sheet['B7'].value 
            f1 = sheet['B8'].value
            if ACC is not None:
                self.accuracyLabel.setText('Accuracy = %.4f' % ACC)
            if SEN is not None:
                self.sensitivityLabel.setText('Sensitivity = %.4f' % SEN)
            if SPE is not None: 
                self.specificityLabel.setText('Specificity = %.4f' % SPE)
            if f1 is not None:
                self.f1Label.setText('F1-score = %.4f' % f1)        
        if self.radioButtonLine.isChecked():
            self.graphWidget1.plot(data["rrTsVec"], data["rrVec"], pen='b')
        else:
            self.graphWidget1.plot(data["rrTsVec"], data["rrVec"], pen=None, symbol='o', symbolSize=5)
        self.graphWidget2.plot(data["estAfTsVec"], data["estAfVec"], pen="g")
        self.graphWidget3.plot(data["estAfTsVec"], data["estAfVec"], pen="g")
        hLine1 = pg.InfiniteLine(angle=0, pos=self.threshold.value(), movable=False, pen='r', label='Threshold = %.2f' % self.threshold.value(), labelOpts={'position': 0.9, 'movable': True})
        hLine2 = pg.InfiniteLine(angle=0, pos=self.threshold.value(), movable=False, pen='r', label='Threshold = %.2f' % self.threshold.value(), labelOpts={'position': 0.9, 'movable': True})
        self.graphWidget2.addItem(hLine1, ignoreBounds=True)
        self.graphWidget3.addItem(hLine2, ignoreBounds=True)

        self.graphWidget2.setYRange(-0.1, 1.1, padding=0)
        self.graphWidget3.setYRange(-0.1, 1.1, padding=0)

        self.graphWidget4.setMouseEnabled(x=True, y=True)
        self.graphWidget1.setMouseEnabled(x=True, y=True)
        self.graphWidget2.setMouseEnabled(x=True, y=False)
        self.graphWidget3.setMouseEnabled(x=True, y=False)

        self.vLine4 = pg.InfiniteLine(angle=90, movable=False)
        self.graphWidget4.addItem(self.vLine4)
        self.vLine1 = pg.InfiniteLine(angle=90, movable=False)
        self.graphWidget1.addItem(self.vLine1)
        self.vLine2 = pg.InfiniteLine(angle=90, movable=False)
        self.graphWidget2.addItem(self.vLine2)
        self.region = pg.LinearRegionItem()
        self.region.setZValue(10)
        self.graphWidget3.addItem(self.region, ignoreBounds=True)
        self.graphWidget1.sigRangeChanged.connect(self.updateRegion)
        self.graphWidget2.sigRangeChanged.connect(self.updateRegion)
        self.graphWidget4.sigRangeChanged.connect(self.updateRegion)
        self.region.sigRegionChanged.connect(self.update)
        self.region.setRegion(data["defaultRegion"])
        self.vb4 = self.graphWidget4.plotItem.vb
        self.vb1 = self.graphWidget1.plotItem.vb
        self.vb2 = self.graphWidget2.plotItem.vb
        self.graphWidget4.scene().sigMouseMoved.connect(self.mouseMoved4)
        self.graphWidget1.scene().sigMouseMoved.connect(self.mouseMoved1)
        self.graphWidget2.scene().sigMouseMoved.connect(self.mouseMoved2)
        self.addAfButton.setEnabled(True)
        self.removeAfButton.setEnabled(True)
        self.addEcgAfButton.setEnabled(True)
        self.removeEcgAfButton.setEnabled(True)
        self.labelAllAfRegionsButton.setEnabled(True)
        self.resultAnalysisButton.setEnabled(True)
        self.ecgPlotButton.setEnabled(True)
        color = (255, 0, 0, 50)
        rrRegionPositions = excelProcessing.getRegion(self.workbook, 'Estimated AF Regions')
        if rrRegionPositions is not None:
            self.rrAfRegions.setRegions(self.rrAfRegionTable, self.graphWidget1, 
                                 self.onAfChange, rrRegionPositions, color)
        color = (255, 255, 0, 50)
        ecgRegionPositions = excelProcessing.getRegion(self.workbook, 'Expert AF Regions')
        if ecgRegionPositions is not None:
            self.ecgAfRegions.setRegions(self.ecgAfRegionTable, self.graphWidget4, 
                                 self.onEcgAfChange, ecgRegionPositions, color)        
                
    def onAfChange(self, idx=0, region=None):
        pos = region.getRegion()
        startTime = datetime.fromtimestamp(pos[0])
        startItem = QtWidgets.QTableWidgetItem(
            datetime.strftime(startTime, '%Y-%m-%d %H:%M:%S.%f'))
        endTime = datetime.fromtimestamp(pos[1])
        endItem = QtWidgets.QTableWidgetItem(
            datetime.strftime(endTime, '%Y-%m-%d %H:%M:%S.%f'))
        self.rrAfRegionTable.setItem(self.rrAfRegions.regionList.index(idx), 1, startItem)
        self.rrAfRegionTable.setItem(self.rrAfRegions.regionList.index(idx), 2, endItem)
        
    def onEcgAfChange(self, idx=0, region=None):
        pos = region.getRegion()
        startTime = datetime.fromtimestamp(pos[0])
        startItem = QtWidgets.QTableWidgetItem(
            datetime.strftime(startTime, '%Y-%m-%d %H:%M:%S.%f'))
        endTime = datetime.fromtimestamp(pos[1])
        endItem = QtWidgets.QTableWidgetItem(
            datetime.strftime(endTime, '%Y-%m-%d %H:%M:%S.%f'))
        self.ecgAfRegionTable.setItem(self.ecgAfRegions.regionList.index(idx), 1, startItem)
        self.ecgAfRegionTable.setItem(self.ecgAfRegions.regionList.index(idx), 2, endItem)

def main():
    app = QtWidgets.QApplication(sys.argv)
    main = MainWindow()
    main.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
